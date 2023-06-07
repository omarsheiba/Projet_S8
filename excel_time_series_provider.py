import os
from typing import Sequence
from openpyxl import load_workbook

#handles extraction of a time series from an excel workbook
class ExcelTimeSeriesProvider:
    def __init__(self, filepath: str):
        if not os.path.exists(filepath):
            raise OSError(f"market data source file not found {filepath}")
        self.wb_obj = load_workbook(filepath)
        self.nb_sheets = len(self.wb_obj.sheetnames)

#we do not know what design constraints these excel files follows, the fewest assumption regarding their design have been made
#basically we only assume that the time series are vertical non discontinued ranges (could be located anywhere on any of the spreadsheets of the workbook)
#and labels describing the timeseries are vertically disposed above it
#this will still work if you move the ranges around in the spreadsheet
#or if the workbook contains multiple spreadsheets
    def get_time_series(self, *args) -> Sequence[str]:
        i=0
        res = []
        args = list(args)
        while i < self.nb_sheets:
            j=1
            active_sheet = self.wb_obj[self.wb_obj.sheetnames[i]]
            while j <= active_sheet.max_row:
                k=1
                while k <= active_sheet.max_column:
                    label_range_start = chr(ord('@')+k) + str(j)
                    label_range_end = chr(ord('@') +k) + str(j+len(args)-1)
                    if [v[0].value for v in active_sheet[f"{label_range_start}:{label_range_end}"]] == args:
                        for l in range(j+len(args), active_sheet.max_row +1):
                            if active_sheet.cell(l,k) is None:
                                print(f"Warning : cell {l,k} in time series {args} is empty")
                            res.append(active_sheet.cell(l,k).value)
                        return res
                    k+=1
                j+=1
            i+=1 
       